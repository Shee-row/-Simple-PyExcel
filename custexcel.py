from openpyxl import Workbook, load_workbook


wb = Workbook()
# Name worksheet that is currently active in the workbook
ws = wb.active
# Naming our Excel
wb.title = "Sales"
# Putting entries in our Excel
ws.append(["Cust_ID", "Cust_Name", "Item", "Unit_Price", "Quantity"])
ws.append([101, "Jane", "Chair", 2500, 1])
ws.append([102, "John", "Table", 4500, 1])
ws.append([103, "Mary", "Bed", 5000, 2])
ws.append([104, "Kim", "Cupboard", 5500, 1])
ws.append([105, "Dennis", "Bed", 5000, 1])
ws.append([106, "Sam", "Chair", 2500, 4])

# To access the Excel file
wbk = load_workbook("Sales2.xlsx")

# To insert new columns
ws.insert_cols(idx=6, amount=2)
ws.cell(1,6).value = "Total"
ws.cell(1,7).value = "Final_Total"

# Passing rows and columns as arguments


def cell_value(r, c):
    return ws.cell(row=r, column=c).value

# For each row present ,calculate the total and store result in the "Total" column


for each_row in range(2, ws.max_column + 1):
    ws.cell(row=each_row, column=6).value = cell_value(each_row, 4) * cell_value(each_row,5)
# Assuming discount is 20% on the total price
    ws.cell(row=each_row, column=7).value = cell_value(each_row, 6) * 0.8

# Save changes made
wb.save("Sales2.xlsx")


